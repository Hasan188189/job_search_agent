"""
Tracker
=======
Persists job data in a local SQLite database and exports to Excel.
"""

import sqlite3
from datetime import datetime
from pathlib import Path
from config.config_loader import load_config
from core.logger import get_logger

logger = get_logger("tracker")


class Tracker:
    def __init__(self):
        cfg = load_config()
        db_path = Path(cfg["tracker"]["db_path"])
        db_path.parent.mkdir(parents=True, exist_ok=True)
        self.conn = sqlite3.connect(str(db_path), check_same_thread=False)
        self._init_db()

    def _init_db(self):
        self.conn.execute("""
            CREATE TABLE IF NOT EXISTS applications (
                id          TEXT PRIMARY KEY,
                platform    TEXT,
                title       TEXT,
                company     TEXT,
                location    TEXT,
                url         TEXT,
                description TEXT,
                posted_date TEXT,
                match_score REAL,
                status      TEXT DEFAULT 'discovered',
                notes       TEXT,
                updated_at  TEXT
            )
        """)
        # Add columns if upgrading from old schema
        for col, coltype in [("description", "TEXT"), ("posted_date", "TEXT")]:
            try:
                self.conn.execute(f"ALTER TABLE applications ADD COLUMN {col} {coltype}")
            except Exception:
                pass
        self.conn.commit()

    def upsert_job(self, job: dict):
        now = datetime.utcnow().isoformat()
        self.conn.execute("""
            INSERT INTO applications
                (id, platform, title, company, location, url, description, posted_date,
                 match_score, status, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 'discovered', ?)
            ON CONFLICT(id) DO UPDATE SET
                match_score = excluded.match_score,
                description = COALESCE(excluded.description, applications.description),
                updated_at  = excluded.updated_at
        """, (
            job["id"], job.get("platform"), job.get("title"), job.get("company"),
            job.get("location"), job.get("url"), job.get("description", ""),
            job.get("posted_date", ""), job.get("match_score", 0), now,
        ))
        self.conn.commit()

    def get_job(self, job_id: str) -> dict | None:
        row = self.conn.execute(
            "SELECT * FROM applications WHERE id = ?", (job_id,)
        ).fetchone()
        return self._row_to_dict(row) if row else None

    def print_dashboard(self):
        total = self.conn.execute("SELECT COUNT(*) FROM applications").fetchone()[0]
        platforms = self.conn.execute(
            "SELECT platform, COUNT(*) FROM applications GROUP BY platform"
        ).fetchall()

        print("\n" + "=" * 55)
        print("  JOB SEARCH RESULTS")
        print("=" * 55)
        for plat, count in platforms:
            print(f"  {plat:<14} {'#' * count:<20} {count:>3}")
        print("-" * 55)
        print(f"  {'TOTAL':<14} {'':20} {total:>3}")
        print("=" * 55)

        recent = self.conn.execute("""
            SELECT title, company, location, platform, match_score
            FROM applications ORDER BY updated_at DESC LIMIT 10
        """).fetchall()
        if recent:
            print("\n  Recent jobs found:")
            for r in recent:
                score = f"{r[4]:.0%}" if r[4] else "N/A"
                print(f"  {r[0][:35]:<35} @ {r[1][:20]:<20} ({r[3]}) {score}")
        print()

    def export_spreadsheet(self, output_path: str = None) -> str:
        """Export all jobs to a well-formatted Excel spreadsheet."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        if not output_path:
            from config.config_loader import PROJECT_ROOT
            output_path = str(PROJECT_ROOT / "data" / "applications.xlsx")

        cols = ["S.No", "Platform", "Job Title", "Company", "Location",
                "Description", "Match Score", "Job URL", "Found On"]
        db_cols = ["platform", "title", "company", "location",
                   "description", "match_score", "url", "updated_at"]

        rows = self.conn.execute(f"""
            SELECT {', '.join(db_cols)} FROM applications
            ORDER BY match_score DESC, updated_at DESC
        """).fetchall()

        wb = Workbook()
        ws = wb.active
        ws.title = "Job Listings"

        # Styles
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        # Headers
        for col_idx, col_name in enumerate(cols, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        # Data rows
        for row_idx, row in enumerate(rows, 2):
            ws.cell(row=row_idx, column=1, value=row_idx - 1).border = thin_border  # S.No
            for col_idx, value in enumerate(row, 2):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if cols[col_idx - 1] == "Match Score" and value is not None:
                    cell.number_format = "0%"
                    cell.alignment = Alignment(horizontal="center", vertical="top")

        # Column widths
        widths = {"A": 6, "B": 12, "C": 40, "D": 25, "E": 25,
                  "F": 50, "G": 12, "H": 55, "I": 20}
        for col_letter, width in widths.items():
            ws.column_dimensions[col_letter].width = width

        # Freeze header row
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        try:
            wb.save(output_path)
        except PermissionError:
            stem = Path(output_path).stem
            alt_path = str(Path(output_path).with_name(f"{stem}_latest.xlsx"))
            wb.save(alt_path)
            logger.info(f"Original file locked, saved to {alt_path}")
            return alt_path
        logger.info(f"Exported {len(rows)} jobs to {output_path}")
        return output_path

    def _row_to_dict(self, row) -> dict:
        cols = ["id", "platform", "title", "company", "location", "url",
                "description", "posted_date", "match_score", "status", "notes",
                "updated_at"]
        return dict(zip(cols, row))
